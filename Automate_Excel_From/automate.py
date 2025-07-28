import os
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Change the current working directory to the directory where this script is located
os.chdir(os.path.dirname(os.path.abspath(__file__)))


def process_workbook(filename):
    # Load the Excel workbook from the given filename
    wb = xl.load_workbook(filename)
    # Select the worksheet named 'Sheet1'
    sheet = wb['Sheet1']

    # Loop through rows starting from row 2 to the last row with data
    # (Assuming row 1 is the header)
    for row in range(2, sheet.max_row + 1):
        # Get the cell in column 3 (C) of the current row
        cell = sheet.cell(row, 3)
        # Calculate corrected price by multiplying the value by 0.9 (apply 10% discount)
        corrected_price = cell.value * 0.9
        # Get the cell in column 4 (D) where corrected price will be stored
        corrected_price_cell = sheet.cell(row, 4)
        # Set the corrected price value into the cell
        corrected_price_cell.value = corrected_price

    # Define the data range for the chart: column 4 (D), from row 2 to last data row
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    # Create a new bar chart object
    chart = BarChart()
    # Add the data range to the chart
    chart.add_data(values)
    # Insert the chart into the wor
