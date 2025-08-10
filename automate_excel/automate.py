import os
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# 將工作目錄切換到此腳本所在的資料夾
os.chdir(os.path.dirname(os.path.abspath(__file__)))


def process_workbook(filename):
    # 載入指定檔名的 Excel 工作簿
    wb = xl.load_workbook(filename)
    # 選取名稱為 'Sheet1' 的工作表
    sheet = wb['Sheet1']

    # 從第2列開始逐列迭代到最後一列（假設第1列是標題）
    for row in range(2, sheet.max_row + 1):
        # 取得該列第3欄（C欄）的儲存格
        cell = sheet.cell(row, 3)
        # 將該儲存格的值乘以0.9，計算折扣後價格（打9折）
        corrected_price = cell.value * 0.9
        # 取得第4欄（D欄）的儲存格，用來放置折扣後價格
        corrected_price_cell = sheet.cell(row, 4)
        # 將折扣後價格寫入儲存格
        corrected_price_cell.value = corrected_price

    # 定義圖表資料範圍：第4欄（D欄）從第2列到最後一列
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    # 建立一個長條圖物件
    chart = BarChart()
    # 將資料範圍加入長條圖
    chart.add_data(values)
    # 將長條圖插入工作表，位置在 E2 儲存格
    sheet.add_chart(chart, "E2")
    # 將修改過的工作簿另存新檔，檔名在原本名稱後加上 "_V01"
    new_filename = filename.replace('.xlsx', '_V01.xlsx')
    wb.save(new_filename)
    print(f"已處理檔案 {filename} 並另存為 {new_filename}")


if __name__ == "__main__":
    # 呼叫函式，處理指定的 Excel 檔案
    process_workbook('transactions.xlsx')
    print("Excel 工作簿處理完成。")
